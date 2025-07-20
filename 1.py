from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
import time
import traceback
import random

def latlng_not_empty(driver):
    spans = driver.find_elements(By.CSS_SELECTOR, "h2.bwoZTb span")
    return bool(spans and spans[0].text.strip() != "")

def main():
    excel_path = "1140720.xlsx"
    sheet_name = "台電電桿位置"

    wb = load_workbook(excel_path)
    ws = wb[sheet_name]

    queries = []
    row = 2
    while True:
        val = ws[f"F{row}"].value
        if val is None:
            break
        queries.append((row, str(val)))
        row += 1

    options = Options()
    options.add_argument("--headless")  # 無頭模式
    options.add_argument("--disable-gpu")  # Windows環境建議
    options.add_argument("--window-size=1920,1080")  # 設定視窗大小

    driver = webdriver.Chrome(options=options)
    wait = WebDriverWait(driver, 15)

    driver.get("https://linspace.somee.com/TPCToMap/?Branch=Miaoli")

    for row, q in queries:
        try:
            input_box = wait.until(EC.presence_of_element_located((By.ID, "txtInput")))
            input_box.send_keys(Keys.CONTROL + "a")
            input_box.send_keys(Keys.DELETE)
            time.sleep(0.2)
            input_box.send_keys(q)

            btn = wait.until(EC.element_to_be_clickable((By.ID, "btnLocate")))
            btn.click()

            wait.until(latlng_not_empty)

            latlng_text = driver.find_element(By.CSS_SELECTOR, "h2.bwoZTb span").text.strip()
            lat, lng = [s.strip() for s in latlng_text.split(",")]

            print(f"第{row}列 {q} → 緯度：{lat}, 經度：{lng}")

            ws[f"I{row}"] = lat
            ws[f"J{row}"] = lng
            wb.save(excel_path)  # 每筆存檔防止資料遺失

            time.sleep(random.uniform(1, 2))

            driver.back()
            wait.until(EC.presence_of_element_located((By.ID, "txtInput")))

        except Exception as e:
            print(f"第{row}列 {q} 查詢失敗：{e}")
            traceback.print_exc()
            ws[f"I{row}"] = "查詢失敗"
            ws[f"J{row}"] = "查詢失敗"
            wb.save(excel_path)
            try:
                driver.back()
                wait.until(EC.presence_of_element_located((By.ID, "txtInput")))
            except:
                pass

    driver.quit()
    print("已將結果寫回 Excel 並存檔。")

if __name__ == "__main__":
    main()
